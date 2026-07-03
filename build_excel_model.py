"""
Builds payment_cycle_model.xlsx from invoices_sample.csv.
Sheets: Dashboard, Invoice Ledger, Aging Analysis
Uses live Excel formulas (not hardcoded Python-computed values).
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
YELLOW = PatternFill("solid", start_color="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

df = pd.read_csv("invoices_sample.csv", parse_dates=["issue_date", "due_date", "paid_date"])
n = len(df)

wb = Workbook()

# ---------------- Invoice Ledger ----------------
ledger = wb.active
ledger.title = "Invoice Ledger"
headers = ["Invoice ID", "Customer", "Issue Date", "Due Date", "Terms (days)",
           "Amount ($)", "Paid Date", "Status", "Days to Pay", "Days Past Due", "Aging Bucket"]
for c, h in enumerate(headers, 1):
    cell = ledger.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

for i, row in df.iterrows():
    r = i + 2
    ledger.cell(row=r, column=1, value=row["invoice_id"]).font = BLUE
    ledger.cell(row=r, column=2, value=row["customer"]).font = BLUE
    ledger.cell(row=r, column=3, value=row["issue_date"].strftime("%Y-%m-%d")).font = BLUE
    ledger.cell(row=r, column=4, value=row["due_date"].strftime("%Y-%m-%d")).font = BLUE
    ledger.cell(row=r, column=5, value=int(row["terms_days"])).font = BLUE
    ledger.cell(row=r, column=6, value=round(float(row["amount"]), 2)).font = BLUE
    paid_val = row["paid_date"].strftime("%Y-%m-%d") if pd.notna(row["paid_date"]) else ""
    ledger.cell(row=r, column=7, value=paid_val).font = BLUE
    ledger.cell(row=r, column=8, value=f'=IF(G{r}="","Open","Paid")').font = BLACK
    ledger.cell(row=r, column=9, value=f'=IF(G{r}="","",DATEVALUE(G{r})-DATEVALUE(C{r}))').font = BLACK
    ledger.cell(row=r, column=10,
                value=f'=IF(G{r}="",TODAY()-DATEVALUE(D{r}),DATEVALUE(G{r})-DATEVALUE(D{r}))').font = BLACK
    ledger.cell(row=r, column=11,
                value=(f'=IF(J{r}<=0,"Current",IF(J{r}<=30,"1-30 days",'
                       f'IF(J{r}<=60,"31-60 days",IF(J{r}<=90,"61-90 days","90+ days"))))')).font = BLACK
    for c in range(1, 12):
        ledger.cell(row=r, column=c).border = BORDER

ledger.cell(row=1, column=6).comment = None
for col, width in zip("ABCDEFGHIJK", [12, 14, 12, 12, 12, 12, 12, 10, 12, 13, 13]):
    ledger.column_dimensions[col].width = width
ledger.freeze_panes = "A2"

# ---------------- Aging Analysis ----------------
aging = wb.create_sheet("Aging Analysis")
aging["A1"] = "Open AR — Aging Analysis"
aging["A1"].font = TITLE_FONT
buckets = ["Current", "1-30 days", "31-60 days", "61-90 days", "90+ days"]
aging["A3"] = "Bucket"
aging["B3"] = "Open Amount ($)"
aging["C3"] = "% of Open AR"
for c in ("A3", "B3", "C3"):
    aging[c].font = HEADER_FONT
    aging[c].fill = HEADER_FILL

last_row = n + 1
for i, b in enumerate(buckets):
    r = 4 + i
    aging.cell(row=r, column=1, value=b).font = BLACK
    formula = (f'=SUMIFS(\'Invoice Ledger\'!F2:F{last_row},'
               f'\'Invoice Ledger\'!H2:H{last_row},"Open",'
               f'\'Invoice Ledger\'!K2:K{last_row},A{r})')
    aging.cell(row=r, column=2, value=formula).font = BLACK
    aging.cell(row=r, column=2).number_format = '$#,##0;($#,##0);-'
    aging.cell(row=r, column=3, value=f'=B{r}/SUM($B$4:$B$8)').font = BLACK
    aging.cell(row=r, column=3).number_format = "0.0%"

aging["A9"] = "Total Open AR"
aging["A9"].font = Font(bold=True)
aging["B9"] = "=SUM(B4:B8)"
aging["B9"].font = Font(bold=True)
aging["B9"].number_format = '$#,##0;($#,##0);-'

chart = BarChart()
chart.title = "Open AR by Aging Bucket"
chart.y_axis.title = "Amount ($)"
data = Reference(aging, min_col=2, min_row=3, max_row=8)
cats = Reference(aging, min_col=1, min_row=4, max_row=8)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width, chart.height = 16, 9
aging.add_chart(chart, "E3")
for col, width in zip("ABC", [16, 18, 14]):
    aging.column_dimensions[col].width = width

# ---------------- Dashboard ----------------
dash = wb.create_sheet("Dashboard", 0)
dash["A1"] = "Payment Cycle Dashboard"
dash["A1"].font = TITLE_FONT
dash["A2"] = "All figures calculated live from Invoice Ledger"
dash["A2"].font = Font(italic=True, size=9, color="808080")

metrics = [
    ("Total Invoices", f'=COUNTA(\'Invoice Ledger\'!A2:A{last_row})', "0"),
    ("Total Billed ($)", f"=SUM('Invoice Ledger'!F2:F{last_row})", '$#,##0;($#,##0);-'),
    ("Open AR Balance ($)", f"=SUMIFS('Invoice Ledger'!F2:F{last_row},'Invoice Ledger'!H2:H{last_row},\"Open\")", '$#,##0;($#,##0);-'),
    ("Avg Days to Pay", f"=AVERAGEIFS('Invoice Ledger'!I2:I{last_row},'Invoice Ledger'!H2:H{last_row},\"Paid\")", "0.0"),
    ("% Paid On Time", f"=COUNTIFS('Invoice Ledger'!H2:H{last_row},\"Paid\",'Invoice Ledger'!J2:J{last_row},\"<=0\")/COUNTIFS('Invoice Ledger'!H2:H{last_row},\"Paid\")", "0.0%"),
    ("% Open AR 90+ Days", "='Aging Analysis'!C8", "0.0%"),
]
r0 = 4
dash.cell(row=r0, column=1, value="Metric").font = HEADER_FONT
dash.cell(row=r0, column=1).fill = HEADER_FILL
dash.cell(row=r0, column=2, value="Value").font = HEADER_FONT
dash.cell(row=r0, column=2).fill = HEADER_FILL
for i, (label, formula, fmt) in enumerate(metrics):
    r = r0 + 1 + i
    dash.cell(row=r, column=1, value=label).font = BLACK
    cell = dash.cell(row=r, column=2, value=formula)
    cell.font = Font(bold=True, color="1F4E78")
    cell.number_format = fmt
    if "90+" in label or "On Time" in label:
        dash.cell(row=r, column=2).fill = YELLOW

dash["A12"] = "Note: DSO target = 45 days. See README for methodology."
dash["A12"].font = Font(italic=True, size=9, color="808080")
dash.column_dimensions["A"].width = 24
dash.column_dimensions["B"].width = 20

wb.save("payment_cycle_model.xlsx")
print("Saved payment_cycle_model.xlsx")
